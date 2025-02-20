from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import os
from datetime import datetime
from selenium.common.exceptions import NoSuchElementException, TimeoutException

# Define the URL of the AlgoThink store
store_url = 'https://www.teacherspayteachers.com/store/storename'

# ------------------------------------------------------------------------------
# Function to scrape product details directly from the main page
# ------------------------------------------------------------------------------
def scrape_products_on_main_page(driver):
    """
    Navigates through all pages of the store and scrapes:
      - Product Title
      - Price
      - Grade Levels
      - Product Link
    Returns a list of dictionaries with the above information.
    """
    scraped_data = []
    product_number = 0  # Counter for product numbers

    while True:
        try:
            # Wait for products to load
            WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, '.ProductRowCard-module__card--xTOd6'))
            )
            print("Fetching product details from the current page...")

            # Select all product cards using the appropriate class
            products = driver.find_elements(By.CSS_SELECTOR, '.ProductRowCard-module__card--xTOd6')

            for product in products:
                # Scrape product title
                try:
                    title = product.find_element(By.CSS_SELECTOR, 'h2 a').text
                except:
                    title = 'N/A'

                # Scrape price
                try:
                    price_element = product.find_element(
                        By.CSS_SELECTOR, '[data-testid="stacked-sale-or-discount-price"]'
                    )
                    price = price_element.text
                except NoSuchElementException:
                    price = 'N/A'

                # Scrape grade levels
                try:
                    grade_levels = product.find_element(
                        By.CSS_SELECTOR, '.MetadataFacetSection__row .Text-module__root--Jk_wf'
                    ).text
                except:
                    grade_levels = 'N/A'

                # Capture the product link
                try:
                    product_link = product.find_element(By.CSS_SELECTOR, 'h2 a').get_attribute('href')
                except:
                    product_link = 'N/A'

                # Store product data in a dictionary
                product_info = {
                    'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'Product Number': 0,  # Placeholder for now
                    'Product Title': title,
                    'Product Price': price,
                    'Grade Levels': grade_levels,
                    'Product Link': product_link
                }

                scraped_data.append(product_info)
                product_number += 1  # Increment the product counter

            # Attempt to find and click the 'Next' button to go to the next page
            try:
                next_button = driver.find_element(By.CSS_SELECTOR, '[data-testid="next-button"]')
                if next_button.is_enabled():
                    next_button.click()
                    WebDriverWait(driver, 10).until(EC.staleness_of(products[0]))
                else:
                    break  # If next button is not enabled, exit loop
            except (NoSuchElementException, TimeoutException):
                break  # Exit loop if the next button is not found or navigation fails

        except Exception as e:
            print(f"ERROR: Failed to retrieve product details from the page: {e}")
            break

    # Reverse product numbers so the first product scraped is the largest number
    for i, product in enumerate(reversed(scraped_data), start=1):
        product['Product Number'] = i

    return scraped_data

# ------------------------------------------------------------------------------
# Function to save scraped data to a new Excel spreadsheet
# ------------------------------------------------------------------------------
def save_to_numbers(data, workbook_path):
    """
    Saves the scraped data to an Excel file (.xlsx). 
    - Creates a new file if none exists.
    - Adds new price columns (shifting older prices to the right).
    - Updates existing products with the latest price data.
    """
    try:
        print("Saving scraped data to spreadsheet...")

        # Define the headers to be used
        headers = ['Timestamp', 'Product Number', 'Product Title', 'Grade Levels', 'Product Link']

        # Check if the file exists; create a new one if it doesn't
        if not os.path.exists(workbook_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Product Data'

            # Write headers to the first row
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num).value = header

            workbook.save(workbook_path)
            print(f"Spreadsheet created at {workbook_path}")
        else:
            # Load the existing workbook
            workbook = openpyxl.load_workbook(workbook_path)
            sheet = workbook.active

            # Check if headers are present; if not, add them
            current_headers = [cell.value for cell in sheet[1]]
            if current_headers != headers:
                for col_num, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=col_num).value = header

        # Dictionary to map product titles to their row numbers
        existing_products = {
            sheet.cell(row=row, column=3).value: row
            for row in range(2, sheet.max_row + 1)
        }

        # Shift price data to the right to make room for the latest prices
        for row in range(2, sheet.max_row + 1):
            for col in range(sheet.max_column, 5, -1):
                sheet.cell(row=row, column=col + 1).value = sheet.cell(row=row, column=col).value

        # Add new price column header with the current date
        sheet.cell(row=1, column=6).value = f'Price {datetime.now().strftime("%Y-%m-%d")}'

        # Populate / Update data
        for product in data:
            timestamp = product['Timestamp']
            title = product['Product Title']
            price = product['Product Price']
            product_number = product['Product Number']
            grade_levels = product['Grade Levels']
            product_link = product['Product Link']

            # Check if product already exists in the spreadsheet
            if title in existing_products:
                row = existing_products[title]
                sheet.cell(row=row, column=2).value = product_number
                sheet.cell(row=row, column=6).value = price
                sheet.cell(row=row, column=5).value = product_link
            else:
                next_row = sheet.max_row + 1
                sheet.cell(row=next_row, column=1).value = timestamp
                sheet.cell(row=next_row, column=2).value = product_number
                sheet.cell(row=next_row, column=3).value = title
                sheet.cell(row=next_row, column=4).value = grade_levels
                sheet.cell(row=next_row, column=5).value = product_link
                sheet.cell(row=next_row, column=6).value = price

                # Update the dictionary with the new product's row number
                existing_products[title] = next_row

        workbook.save(workbook_path)
        print(f"Data saved successfully to {workbook_path}.")

    except Exception as e:
        print(f"ERROR: An error occurred while saving data to the spreadsheet: {e}")

# ------------------------------------------------------------------------------
# Main Execution
# ------------------------------------------------------------------------------
if __name__ == "__main__":
    print("TPT-SCRAPE-3 ACTIVATING...")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    try:
        driver.get(store_url)

        # Scrape the products from all pages
        scraped_data = scrape_products_on_main_page(driver)

        if scraped_data:
            # Save to a dedicated spreadsheet
            save_to_numbers(scraped_data, 'Spider-2-Data.xlsx')
        else:
            print("No product data found on the main page.")

    finally:
        driver.quit()
        print("TPT-Spider-2 COMPLETE")
