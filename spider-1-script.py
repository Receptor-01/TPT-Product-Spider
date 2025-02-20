from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import os
from datetime import datetime
import time

# ------------------------------------------------------------------------------
# URL of your Teachers Pay Teachers store
# ------------------------------------------------------------------------------
store_url = 'https://www.teacherspayteachers.com/store/storename'

# ------------------------------------------------------------------------------
# Function to scrape key store information using Selenium
# ------------------------------------------------------------------------------
def scrape_tpt_store(url):
    """
    Scrapes the following data from the given TPT store page:
      - 5-Star Rating
      - Number of Reviews
      - Number of Followers
      - Number of Products
    Returns a dictionary with the scraped data.
    """
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    try:
        # Open the TPT store page
        driver.get(url)
        time.sleep(5)  # Allow some time for the page to fully load

        # Scrape the 5-star rating
        try:
            rating_element = driver.find_element(
                By.CSS_SELECTOR, 
                '.RatingsLabel-module__singleLineLabel--J9KuO'
            )
            rating = rating_element.text.split()[0]
        except:
            rating = 'N/A'

        # Scrape the number of reviews
        try:
            reviews_count_element = driver.find_element(
                By.CSS_SELECTOR, 
                '.RatingsLabel-module__totalReviews--Roe3y'
            )
            reviews_count = reviews_count_element.text.strip("()")
        except:
            reviews_count = 'N/A'

        # Scrape the number of followers
        try:
            followers_element = driver.find_element(
                By.XPATH, 
                '/html/body/div[1]/div/div/div[2]/div[1]/div[2]/div/div/div[2]/div[1]/div[1]/div[2]'
            )
            followers = followers_element.text.split()[0]
        except:
            followers = 'N/A'

        # Scrape the total number of products
        try:
            products_count_element = driver.find_element(
                By.CSS_SELECTOR, 
                '.StorePageCategoriesList-module__count--DCtKc'
            )
            products_count = products_count_element.text
        except:
            products_count = 'N/A'

        # Return collected data
        return {
            '5-Star Rating': rating,
            'Number of Reviews': reviews_count,
            'Number of Followers': followers,
            'Number of Products': products_count,
        }

    except Exception as e:
        print(f"ERROR: An error occurred while scraping the store: {e}")
        return {}
    finally:
        driver.quit()

# ------------------------------------------------------------------------------
# Function to save scraped data to an Excel spreadsheet
# ------------------------------------------------------------------------------
def save_to_numbers(data):
    """
    Saves store data (rating, reviews, followers, products) to an Excel file 
    called 'Spider-1-Data.xlsx'. Automatically creates the file if it doesn't exist.
    Appends a new row with a timestamp and the latest scraped information.
    """
    try:
        # Define the new workbook path
        workbook_path = 'Spider-1-Data.xlsx'

        # Prepare headers
        headers = ['Timestamp'] + list(data.keys())

        # Check if the file exists; create new if needed
        if not os.path.exists(workbook_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'TPT Store Data'

            # Write headers to the first row
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num).value = header
        else:
            workbook = openpyxl.load_workbook(workbook_path)
            sheet = workbook.active
            # Check if headers are present; if not, add them
            current_headers = [cell.value for cell in sheet[1]]
            if current_headers != headers:
                for col_num, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=col_num).value = header

        # Find the next empty row
        next_row = sheet.max_row + 1
        # Timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row=next_row, column=1).value = timestamp

        # Fill data in columns 2 onwards
        for col_num, key in enumerate(data.keys(), 2):
            value = ', '.join(data[key]) if isinstance(data[key], list) else data[key]
            sheet.cell(row=next_row, column=col_num).value = value

        # Save the workbook
        workbook.save(workbook_path)

    except Exception as e:
        print(f"ERROR: An error occurred while saving data to the spreadsheet: {e}")

# ------------------------------------------------------------------------------
# Main Execution
# ------------------------------------------------------------------------------
if __name__ == "__main__":
    # Scrape data from your store
    scraped_data = scrape_tpt_store(store_url)

    # If we got any data back, save it to Excel
    if scraped_data:
        save_to_numbers(scraped_data)
